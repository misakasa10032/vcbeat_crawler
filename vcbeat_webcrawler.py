# -*- coding: utf-8 -*-
import requests
import json
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

num_items = input('Please input the number of items that you would like to search for: ')
num_page = math.ceil(int(num_items)/10)
url_0 = 'https://vcbeat.cn/data/comp_search'	#	displaying the general summary of items
headers = {'content-type': 'application/json'}
payload_0 = {"query":{"round_id":"","demostic":"","xxfly_id":""},"page":2}
wb = Workbook()
sheet = wb['Sheet']
count_m = 1
dict_flag = {}
a0 = input('Short Name(Y/N): ')
a1 =  input('Full Name(Y/N): ')
a2 = input('Address in English & Chinese(Y/N): ')
a3 = input('Brief Introduction(Y/N): ')
a4 = input('General Category(Y/N): ')
a5 = input('E-mail Address(Y/N): ')
a6 = input('Tel Number(Y/N): ')
a7 = input('Founding Time(Y/N): ')
a8 = input('Whether Domestic or Not(Y/N): ')
a9 = input('Running State(Y/N): ')
a10 = input('Count of Employees(Y/N): ')
a11 = input('Events of raising funds(Y/N): ')
dict_flag['Short Name'] = a0
dict_flag['Full Name'] = a1
dict_flag['Address in English & Chinese'] = a2
dict_flag['Brief Introduction'] = a3
dict_flag['General Category'] = a4
dict_flag['E-mail Address'] = a5
dict_flag['Tel Number'] = a6
dict_flag['Founding Time'] = a7
dict_flag['Whether Domestic or Not'] = a8
dict_flag['Running State'] = a9
dict_flag['Count of Employees'] = a10
dict_flag['Events of raising funds'] = a11
sum_flag = 0
att_index = {}
for z in list(dict_flag.keys()):
	if dict_flag[z] == 'Y':
		sum_flag += 1
		sheet.cell(row = 1, column = sum_flag).value = z
		att_index[sum_flag] = z
for k in range(1, num_page + 1):
	payload_0['page'] = k
	r = requests.post(url_0, data = json.dumps(payload_0), headers = headers)
	j_r = r.content.decode()
	dict_j0 = json.loads(j_r)
	data_list = dict_j0['data']
	for item in data_list:
		count_m += 1
		uid = item['uid']	#	uid code
		if not item['short_name'] is None:
			sn = item['short_name']	#	Short Name
		else:
			sn = 'not released'
		if not item['full_name'] is None:
			fn = item['full_name']	#	Full Name
		else:
			fn = 'not released'
		if (not item['province'] is None and not item['city'] is None):
			place = item['province'] + '-' + item['city']	#	venue of its office in Chinese
		elif not item['province'] is None:
			place = item['province']
		elif not item['city'] is None:
			place = item['city']
		else:
			place = 'not released'
		if not item['summary'] is None:
			summary = item['summary']	#	Brief Introduction
		else:
			summary = 'not released'
		if (not item['xxfly'] is None and not item['description'] is None):
			category = item['xxfly'] + '-' + item['description']
		elif not item['description'] is None:
			category = item['description']
		elif not item['xxfly'] is None:
			category = item['xxfly']
		else:
			category = 'not released'
		# More information
		url_1 = 'https://vcbeat.cn/data/entity_detail'
		payload_1 = {}
		payload_1['uid'] = uid
		r1 = requests.post(url_1, data = json.dumps(payload_1), headers = headers)
		jr = r1.content.decode()
		dict_j1 = json.loads(jr)
		# Section of base info
		bi = dict_j1['data']['base_info']
		prop_avail2 = list(bi.keys())
		if not bi['address'] is None:
			add = bi['address'] + '/' + place
		else:
			add = place
		if not bi['email'] is None:
			email = bi['email']	#	email address
		else:
			email = 'not released'
		if not bi['tel'] is None:
			tel = bi['tel']	#	Tel Number
		else:
			tel = 'not released'
		if not bi['create_time'] is None:
			create_time = bi['create_time'][0:10]	#	establishment time
		else:
			create_time = 'not released'
		setup_time = bi['setup_time']	#	uploading time
		if not bi['demostic'] is None:
			if bi['demostic'] == '0':
				domestic = '国内'
			else:
				domestic = '国外'	#	Whether home or not
		else:
			domestic = 'not released'
		if not bi['state'] is None:
			state = bi['state']	#	operation status
		else:
			state = 'not released'
		if not bi['employee'] is None:
			employee = bi['employee']	# count of employees
		else:
			employee = 'not released'
		# Events of raising funds.
		if 'event' in list(dict_j1['data'].keys()):
			event_list = dict_j1['data']['event']
			rf = ''
			for u in event_list:
				if not u['round_id'] is None:
					round_id = u['round_id']	# round id
					round_dict = {'1':'种子轮', '2':'天使轮', '3':'PreA轮', '4':'A轮', '5':'A+轮', '6':'B轮', '7':'B+轮', '8':'C轮', '9':'C+轮', '10': 'D轮', '11':'PreIPO', '12': 'IPO', '13': '战略投资', '14': '定向增发', '15': '未公开', '16': '捐赠/众筹', '17': '股权转让', '18': '新三板', '20': '债权融资', '21': 'E轮', '22': 'F轮'}
					if round_id in list(round_dict.keys()):
						rounds = round_dict[round_id]
					else:
						rounds = str(round_id)
				else:
					rounds = '被并购'
				if not u['unit_id'] is None:
					unit_dict = {'1': '人民币', '3': '美元', '5': '欧元'}
					if u['unit_id'] in list(unit_dict.keys()):
						unit_name = unit_dict[u['unit_id']]	# the unit of currency
					else:
						unit_name = u['unit_id']
				if not u['amount'] is None:
					amount = u['amount'] + unit_name
				else:
					amount = '(the amount of funds is not released)'
				if not u['precise'] is None:
					precise = u['precise']	#	precision of fund quantity
				else:
					precise = 'not released'
				if not u['event_time'] is None:
					event_time = u['event_time'][0:10]	#	time of raising
				else:
					event_time = '(the time is not released)'
				#	Orgs involved in raising money
				if not u['orgs'] is None:
					orgs_list = u['orgs']
					rf += rounds + ' ' + amount + ' ' + event_time
					for s in orgs_list:
						rc = s['short_name']	#	Names of orgs involved in raising money.
						rf += ' ' + rc 
					rf += '\n'
				else:
					rf += '\n'
			rf= rf.rstrip('\n')
		else:
			rf = 'the information of raising funds is not released'
		attributes = {'Short Name': sn, 'Full Name': fn, 'Address in English & Chinese': add, 'Brief Introduction': summary, 'General Category': category, 'E-mail Address': email, 'Tel Number': tel, 'Founding Time': create_time, 'Whether Domestic or Not': domestic, 'Running State': state, 'Count of Employees': employee, 'Events of raising funds': rf} 
		for j in range(1, 1+ sum_flag):
			sheet.cell(row = count_m, column = j).value = attributes[att_index[j]]
for ro in range(2, sheet.max_row + 1):
	for co in range(1, sheet.max_column + 1):
		sheet.cell(row = ro, column = co).font = Font(name = '等线', size = 11)
		sheet.cell(row = ro, column = co).border = Border(left = Side(border_style = None, color = 'FF000000'), right = Side(border_style = None, color = 'FF000000'), top = Side(border_style = None, color = 'FF000000'), bottom = Side(border_style = None, color = 'FF000000'))
for co in range(1, sheet.max_column + 1):
	sheet.cell(row = 1, column = co).font = Font(name = '等线', bold = True) 
	sheet.cell(row = ro, column = co).border = Border(left = Side(border_style = None, color = 'FF000000'), right = Side(border_style = None, color = 'FF000000'), top = Side(border_style = None, color = 'FF000000'), bottom = Side(border_style = None, color = 'FF000000'))
dict_width = {'Short Name': 28.5, 'Full Name': 35.5, 'Address in English & Chinese': 60, 'Brief Introduction': 80, 'General Category': 39, 'E-mail Address': 27, 'Tel Number': 27, 'Founding Time': 13.5, 'Whether Domestic or Not': 22.5, 'Running State': 12.5, 'Count of Employees': 18.5, 'Events of raising funds': 70}
alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for j in range(1, 1 + sum_flag): 
	sheet.column_dimensions[alphabet[j - 1]].width = dict_width[att_index[j]]
for i in range(2, 1 + sheet.max_row):
	for j in range(1, 1 + sheet.max_column):
		sheet.cell(row = i, column = j).alignment = Alignment(vertical = 'center', wrap_text = True)
wb.save('data_set.xlsx')